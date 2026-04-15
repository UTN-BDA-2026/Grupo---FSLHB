from flask import Blueprint, request, make_response, render_template, jsonify
from app import require_admin, csrf
incidencia_bp = Blueprint('incidencia', __name__)

@incidencia_bp.route('/ranking/tarjetas/excel', methods=['GET'])
def exportar_tarjetas_excel():
    """Genera un archivo Excel con el ranking de tarjetas (amarillas y rojas) para el torneo/categoría/fecha seleccionados."""
    torneo = request.args.get('torneo')
    categoria = request.args.get('categoria')
    fecha = request.args.get('fecha', type=int)

    datos = IncidenciaService.ranking_resumen(torneo, categoria, fecha)
    if (not datos or len(datos) == 0) and torneo:
        datos = IncidenciaService.ranking_resumen(None, categoria, fecha)

    from app.repositories.club_repositorio import ClubRepository
    from app.repositories.jugadora_repositorio import JugadoraRepository

    conteo_verdes = {}
    conteo_amarillas = {}
    conteo_rojas = {}
    for inc in datos:
        # En nuestro modelo, las tarjetas vienen como tipo 'tarjeta' y color 'amarilla' o 'roja'
        if getattr(inc, 'tipo', None) == 'tarjeta':
            club_obj = ClubRepository.buscar_por_id(inc.club_id)
            jug_obj = JugadoraRepository.buscar_por_id(inc.jugadora_id)
            club = club_obj.nombre if club_obj else str(inc.club_id)
            jugador = (f"{jug_obj.nombre} {jug_obj.apellido}".strip()) if jug_obj else str(inc.jugadora_id)
            key = (club, jugador)
            if getattr(inc, 'color', None) == 'verde':
                conteo_verdes[key] = conteo_verdes.get(key, 0) + 1
            elif getattr(inc, 'color', None) == 'amarilla':
                conteo_amarillas[key] = conteo_amarillas.get(key, 0) + 1
            elif getattr(inc, 'color', None) == 'roja':
                conteo_rojas[key] = conteo_rojas.get(key, 0) + 1

    # Unir ambos conteos en una sola lista
    all_keys = set(conteo_verdes.keys()) | set(conteo_amarillas.keys()) | set(conteo_rojas.keys())
    items = [
        {
            'club': k[0],
            'jugador': k[1],
            'verdes': conteo_verdes.get(k, 0),
            'amarillas': conteo_amarillas.get(k, 0),
            'rojas': conteo_rojas.get(k, 0)
        }
        for k in sorted(
            all_keys,
            key=lambda x: (-(conteo_verdes.get(x, 0) + conteo_amarillas.get(x, 0) + conteo_rojas.get(x, 0)), x[1])
        )
    ]

    # Construir XLSX en memoria
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Tarjetas'

    # Encabezado
    ws.append(['Club', 'Jugador', 'Verdes', 'Amarillas', 'Rojas'])
    # Filas
    for it in items:
        ws.append([it['club'], it['jugador'], it['verdes'], it['amarillas'], it['rojas']])

    # Auto-ajustar ancho de columnas básico
    for col in range(1, 6):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            val = row[0].value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 40)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = make_response(bio.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    filename = f"tarjetas_{(torneo or 'torneo').replace(' ', '_')}_{(categoria or 'categoria').replace(' ', '_')}.xlsx"
    resp.headers['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    return resp
from app.services.incidencia_service import IncidenciaService
from app.repositories.partido_repositorio import PartidoRepository
from app.repositories.jugadora_repositorio import JugadoraRepository
from app.repositories.club_repositorio import ClubRepository

# Blueprint ya creado arriba

@incidencia_bp.route('/ranking/goleadores', methods=['GET'])
def ranking_goleadores():
    torneo = request.args.get('torneo')
    categoria = request.args.get('categoria')
    # Buscar partidos filtrados
    filtros = {}
    if torneo:
        filtros['torneo'] = torneo
    if categoria:
        filtros['categoria'] = categoria
    partidos = PartidoRepository.buscar(filtros)
    partido_ids = [p.id for p in partidos]
    # Buscar incidencias de goles
    goles = IncidenciaService.listar_goles_por_partidos(partido_ids)
    ranking = {}
    for inc in goles:
        jugadora = JugadoraRepository.buscar_por_id(inc.jugadora_id)
        club = ClubRepository.buscar_por_id(inc.club_id)
        nombre_jug = ''
        if jugadora:
            nombre_jug = f"{jugadora.nombre} {jugadora.apellido}".strip()
        key = (club.nombre if club else '', nombre_jug)
        ranking[key] = ranking.get(key, 0) + 1
    # Convertir a lista ordenada
    resultado = [
        {'club': k[0], 'jugador': k[1], 'goles': v}
        for k, v in sorted(ranking.items(), key=lambda x: -x[1])
    ]
    return jsonify(resultado)

@incidencia_bp.route('/ranking/tarjetas', methods=['GET'])
def ranking_tarjetas():
    torneo = request.args.get('torneo')
    categoria = request.args.get('categoria')
    color = request.args.get('color')  # 'amarilla' o 'roja'
    filtros = {}
    if torneo:
        filtros['torneo'] = torneo
    if categoria:
        filtros['categoria'] = categoria
    partidos = PartidoRepository.buscar(filtros)
    partido_ids = [p.id for p in partidos]
    tarjetas = IncidenciaService.listar_tarjetas_por_partidos(partido_ids, color)
    ranking = {}
    for inc in tarjetas:
        jugadora = JugadoraRepository.buscar_por_id(inc.jugadora_id)
        club = ClubRepository.buscar_por_id(inc.club_id)
        nombre_jug = ''
        if jugadora:
            nombre_jug = f"{jugadora.nombre} {jugadora.apellido}".strip()
        key = (club.nombre if club else '', nombre_jug)
        ranking[key] = ranking.get(key, 0) + 1
    resultado = [
        {'club': k[0], 'jugador': k[1], 'tarjetas': v}
        for k, v in sorted(ranking.items(), key=lambda x: -x[1])
    ]
    return jsonify(resultado)

@incidencia_bp.route('/ranking/resumen', methods=['GET'])
def ranking_resumen():
    torneo = request.args.get('torneo')
    categoria = request.args.get('categoria')
    fecha = request.args.get('fecha', type=int)
    datos = IncidenciaService.ranking_resumen(torneo, categoria, fecha)
    goleadores = {}
    amarillas = {}
    rojas = {}
    from app.repositories.club_repositorio import ClubRepository
    from app.repositories.jugadora_repositorio import JugadoraRepository
    for row in datos:
        # row es (Incidencia, torneo, categoria)
        inc = row[0] if isinstance(row, (list, tuple)) else row
        club_obj = ClubRepository.buscar_por_id(inc.club_id)
        jug_obj = JugadoraRepository.buscar_por_id(inc.jugadora_id)
        club = club_obj.nombre if club_obj else inc.club_id
        if jug_obj:
            jugador = f"{jug_obj.nombre} {jug_obj.apellido}".strip()
        else:
            jugador = inc.jugadora_id
        key = (club, jugador)
        if inc.tipo == 'gol':
            goleadores[key] = goleadores.get(key, 0) + 1
        elif inc.tipo == 'tarjeta':
            if inc.color == 'amarilla':
                amarillas[key] = amarillas.get(key, 0) + 1
            elif inc.color == 'roja':
                rojas[key] = rojas.get(key, 0) + 1
    def ordenar(dic):
        return [
            {'club': k[0], 'jugador': k[1], 'cantidad': v}
            for k, v in sorted(dic.items(), key=lambda x: -x[1])
        ]
    return jsonify({
        'goleadores': ordenar(goleadores),
        'amarillas': ordenar(amarillas),
        'rojas': ordenar(rojas)
    })


@incidencia_bp.route('/ranking/goleadores/excel', methods=['GET'])
def descargar_goleadores_excel():
    """Genera un Excel (XLSX) con el ranking de goleadores para el torneo/categoría/fecha seleccionados."""
    torneo = request.args.get('torneo')
    categoria = request.args.get('categoria')
    fecha = request.args.get('fecha', type=int)

    # Obtener incidencias y computar conteo de goles por (club, jugador)
    datos = IncidenciaService.ranking_resumen(torneo, categoria, fecha)
    if (not datos or len(datos) == 0) and torneo:
        datos = IncidenciaService.ranking_resumen(None, categoria, fecha)

    from app.repositories.club_repositorio import ClubRepository
    from app.repositories.jugadora_repositorio import JugadoraRepository

    conteo = {}
    for inc in datos:
        if inc.tipo != 'gol':
            continue
        club_obj = ClubRepository.buscar_por_id(inc.club_id)
        jug_obj = JugadoraRepository.buscar_por_id(inc.jugadora_id)
        club = club_obj.nombre if club_obj else str(inc.club_id)
        jugador = (f"{jug_obj.nombre} {jug_obj.apellido}".strip()) if jug_obj else str(inc.jugadora_id)
        key = (club, jugador)
        conteo[key] = conteo.get(key, 0) + 1

    items = [
        {'club': k[0], 'jugador': k[1], 'goles': v}
        for k, v in sorted(conteo.items(), key=lambda x: (-x[1], x[0][1]))
    ]

    # Construir XLSX en memoria
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Goleadores'

    # Encabezado
    ws.append(['Club', 'Jugador', 'Goles'])
    # Filas
    for it in items:
        ws.append([it['club'], it['jugador'], it['goles']])

    # Auto-ajustar ancho de columnas básico
    for col in range(1, 4):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            val = row[0].value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 40)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = make_response(bio.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    filename = f"goleadores_{(torneo or 'torneo').replace(' ', '_')}_{(categoria or 'categoria').replace(' ', '_')}.xlsx"
    resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@incidencia_bp.route('/admin/jugadoras/<int:jugadora_id>/incidencias', methods=['GET'])
@require_admin
def listar_incidencias_por_jugadora_admin(jugadora_id: int):
    """Devuelve las incidencias de tipo tarjeta de una jugadora para uso del panel admin."""
    incidencias = IncidenciaService.listar_tarjetas_por_jugadora(jugadora_id)

    from app.repositories.partido_repositorio import PartidoRepository

    resultado = []
    for inc in incidencias:
        partido = PartidoRepository.buscar_por_id(inc.partido_id)
        resultado.append({
            'id': inc.id,
            'partido_id': inc.partido_id,
            'tipo': inc.tipo,
            'color': getattr(inc, 'color', None),
            'minuto': inc.minuto,
            'torneo': getattr(partido, 'torneo', None) if partido else None,
            'categoria': getattr(partido, 'categoria', None) if partido else None,
            'fecha_numero': getattr(partido, 'fecha_numero', None) if partido else None,
            'club_local_nombre': getattr(getattr(partido, 'club_local', None), 'nombre', None) if partido else None,
            'club_visitante_nombre': getattr(getattr(partido, 'club_visitante', None), 'nombre', None) if partido else None,
        })

    return jsonify(resultado)


@incidencia_bp.route('/admin/jugadoras/<int:jugadora_id>/incidencias', methods=['POST'])
@csrf.exempt
@require_admin
def crear_incidencia_tarjeta_admin(jugadora_id: int):
    """Crea una incidencia de tipo tarjeta para una jugadora desde el panel admin."""
    data = request.get_json(silent=True) or {}
    partido_id = data.get('partido_id')
    club_id = data.get('club_id')
    color = data.get('color')
    minuto = data.get('minuto')

    if not partido_id or not club_id:
        return jsonify({'error': 'partido_id y club_id son obligatorios'}), 400

    if color not in ('verde', 'amarilla', 'roja'):
        return jsonify({'error': 'Color de tarjeta inválido'}), 400

    try:
        inc = IncidenciaService.registrar_tarjeta(partido_id=int(partido_id), club_id=int(club_id), jugadora_id=jugadora_id, color=color, minuto=minuto)
    except Exception as exc:
        return jsonify({'error': 'No se pudo registrar la tarjeta', 'detail': str(exc)}), 500

    return jsonify({'id': inc.id}), 201


@incidencia_bp.route('/admin/incidencias/<int:incidencia_id>', methods=['DELETE'])
@csrf.exempt
@require_admin
def eliminar_incidencia_admin(incidencia_id: int):
    """Elimina una incidencia por id para el panel admin."""
    ok = IncidenciaService.eliminar_por_id(incidencia_id)
    if not ok:
        return jsonify({'error': 'Incidencia no encontrada'}), 404
    return jsonify({'ok': True})
